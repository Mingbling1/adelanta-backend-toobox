"""
⚙️ Motor de Cálculo de Diferido - V2 Arquitectura Hexagonal
Refactorización EXACTA de DiferidoCalcular, DiferidoInternoCalcular y DiferidoExternoCalcular
Mantiene 100% compatibilidad con V1
"""

import pandas as pd
import asyncio
import re
from datetime import datetime, timedelta
from typing import Union
from io import BytesIO
from dateutil.relativedelta import relativedelta
import numpy as np
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import locale

try:
    from utils.adelantafactoring.v2.config.settings import settings
except ImportError:
    # Fallback para desarrollo aislado
    class _FallbackSettings:
        @staticmethod
        def logger(message: str) -> None:
            print(f"[DiferidoCalculationEngine] {message}")

    settings = _FallbackSettings()

# Set locale para compatibilidad V1
try:
    locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")
except Exception:
    pass  # Ignorar si no está disponible

# Diccionario para convertir abreviaturas a nombre completo en español (en minúsculas)
MONTH_MAP = {
    "ENE": "enero",
    "FEB": "febrero",
    "MAR": "marzo",
    "ABR": "abril",
    "MAY": "mayo",
    "JUN": "junio",
    "JUL": "julio",
    "AGO": "agosto",
    "SET": "septiembre",
    "OCT": "octubre",
    "NOV": "noviembre",
    "DIC": "diciembre",
}


def convert_date_col(col: str) -> str:
    """
    Convierte una columna de fecha en el formato '2021AGO' (o variantes en mayúsculas/minúsculas)
    a 'agosto-2021' - LÓGICA EXACTA V1
    """
    m = re.match(r"^(\d{4})([A-Za-z]+)$", col)
    if m:
        year = m.group(1)
        month_abbr = m.group(2).upper()
        month_full = MONTH_MAP.get(month_abbr, month_abbr.lower())
        return f"{month_full}-{year}"
    return col


def reorder_date_columns(date_cols: list[str]) -> list[str]:
    """
    Recibe una lista de columnas de fecha ya convertidas con formato 'mes-YYYY'
    y las retorna ordenadas cronológicamente - LÓGICA EXACTA V1
    """

    def sort_key(col: str):
        try:
            # separa por guión: ej: "enero-2023"
            month_str, year_str = col.split("-")
            year = int(year_str)
            months_order = [
                "enero",
                "febrero",
                "marzo",
                "abril",
                "mayo",
                "junio",
                "julio",
                "agosto",
                "septiembre",
                "octubre",
                "noviembre",
                "diciembre",
            ]
            month_idx = (
                months_order.index(month_str) if month_str in months_order else 99
            )
            return (year, month_idx)
        except Exception:
            return (9999, 99)

    return sorted(date_cols, key=sort_key)


class DiferidoCalculationEngine:
    """
    Motor especializado para cálculos de diferido
    Replica EXACTAMENTE la lógica de DiferidoCalcular, DiferidoInternoCalcular y DiferidoExternoCalcular
    """

    # Constantes de DiferidoInternoCalcular - EXACTO V1
    LISTA_COLUMNAS = [
        "CodigoLiquidacion",
        "NroDocumento",
        "FechaOperacion",
        "FechaConfirmado",
        "Moneda",
        "Interes",
        "DiasEfectivo",
    ]
    DIAS_EFECTIVO = "DiasEfectivo"
    INTERES = "Interes"
    FECHA_PAGO_CONFIRMADO = "FechaConfirmado"
    FECHA_OPERACION = "FechaOperacion"

    def __init__(self):
        self.logger = settings.logger

    def log(self, message: str) -> None:
        """Logging compatible con V1"""
        self.logger(message)

    def last_day_of_month(self, date: datetime) -> datetime:
        """Obtiene el último día del mes - LÓGICA EXACTA V1"""
        if date.month == 12:
            next_month = date.replace(year=date.year + 1, month=1, day=1)
        else:
            next_month = date.replace(month=date.month + 1, day=1)
        return next_month - timedelta(days=1)

    def reorder_date_columns(self, date_cols: list[str]) -> list[str]:
        """
        Recibe una lista de columnas de fecha ya convertidas con formato 'mes-YYYY'
        y las retorna ordenadas cronológicamente - LÓGICA EXACTA V1
        """
        return reorder_date_columns(date_cols)

    # ==================== DIFERIDO INTERNO - LÓGICA EXACTA V1 ====================

    def calcular_monto_por_mes(self, row, mes_inicio, mes_fin):
        """Cálculo de monto por mes - LÓGICA EXACTA DiferidoInternoCalcular V1"""
        dias_total = row[self.DIAS_EFECTIVO]
        monto_total = row[self.INTERES]

        if dias_total <= 0 or np.isnan(dias_total):
            return 0

        if (
            mes_inicio <= row[self.FECHA_PAGO_CONFIRMADO]
            and mes_fin >= row[self.FECHA_OPERACION]
        ):
            mes_inicio_ajustado = max(mes_inicio, row[self.FECHA_OPERACION])
            mes_fin_ajustado = min(mes_fin, row[self.FECHA_PAGO_CONFIRMADO])

            dias_en_mes = (mes_fin_ajustado - mes_inicio_ajustado).days + 1

            porcentaje_dias_mes = dias_en_mes / dias_total

            if porcentaje_dias_mes > 0:
                monto_mes = monto_total * porcentaje_dias_mes
                return monto_mes
            else:
                return 0
        else:
            return 0

    def put_dates_in_columns(
        self,
        df: pd.DataFrame,
        fecha_min: datetime,
        fecha_max: datetime,
    ) -> pd.DataFrame:
        """Poner fechas en columnas - LÓGICA EXACTA DiferidoInternoCalcular V1"""
        fecha_actual = fecha_min
        while fecha_actual <= fecha_max:
            mes_inicio = fecha_actual.replace(day=1)
            mes_fin = self.last_day_of_month(mes_inicio)
            col_name = mes_inicio.strftime("%B-%Y")

            df[col_name] = df.apply(
                lambda row: self.calcular_monto_por_mes(row, mes_inicio, mes_fin),
                axis=1,
            )

            # Avanzar al siguiente mes
            fecha_actual += relativedelta(months=1)
        return df

    def calcular_diferido_interno(
        self, df_interno: pd.DataFrame, hasta: str
    ) -> pd.DataFrame:
        """
        Calcula el diferido interno - LÓGICA EXACTA DiferidoInternoCalcular V1

        Parámetros:
            df_interno (pd.DataFrame): DataFrame interno con datos KPI
            hasta (str): Cadena obligatoria en formato "YYYY-MM"

        Retorna:
            pd.DataFrame: DataFrame con columnas de cada mes calculadas hasta la fecha límite
        """
        if not hasta:
            raise ValueError(
                "El parámetro 'hasta' es obligatorio y debe tener el formato 'YYYY-MM'."
            )

        df = df_interno.copy()
        df = df.loc[:, self.LISTA_COLUMNAS]
        fecha_min = df[self.FECHA_OPERACION].min()
        fecha_max = df[self.FECHA_PAGO_CONFIRMADO].max()
        hasta_date = datetime.strptime(hasta, "%Y-%m")
        hasta_date = self.last_day_of_month(hasta_date)
        df = df[df[self.FECHA_OPERACION] <= hasta_date]
        df = self.put_dates_in_columns(df, fecha_min, fecha_max)
        return df

    # ==================== DIFERIDO EXTERNO - LÓGICA EXACTA V1 ====================

    def read_excel_file(
        self, file_path_or_buffer, sheet_name: str, usecols: str, date_col: str
    ):
        """Leer archivo Excel - LÓGICA EXACTA DiferidoExternoCalcular V1"""
        # Si file_path_or_buffer es un objeto file-like (tiene método 'read')
        if hasattr(file_path_or_buffer, "read"):
            file_path_or_buffer.seek(0)
            df = pd.read_excel(
                file_path_or_buffer, sheet_name=sheet_name, skiprows=1, usecols=usecols
            )
        else:
            # Asumimos que es una ruta de archivo
            df = pd.read_excel(
                file_path_or_buffer, sheet_name=sheet_name, skiprows=1, usecols=usecols
            )
        # Eliminar columnas sin nombre (generadas como 'Unnamed')
        df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
        df[date_col] = pd.to_datetime(df[date_col], dayfirst=True, errors="coerce")
        return df

    def auto_get_usecols(
        self,
        file_path_or_buffer,
        sheet_name: str,
        fixed_range: tuple[str, str],
        stop_marker: str,
    ) -> tuple[str, list[str]]:
        """
        Auto obtener columnas - LÓGICA EXACTA DiferidoExternoCalcular V1
        """
        # Si el archivo está cerrado, reabrirlo usando su nombre (si existe)
        if hasattr(file_path_or_buffer, "closed") and file_path_or_buffer.closed:
            if hasattr(file_path_or_buffer, "name"):
                file_path_or_buffer = open(file_path_or_buffer.name, "rb")
            else:
                raise ValueError("No se puede reabrir el archivo: no tiene 'name'.")
        else:
            if hasattr(file_path_or_buffer, "seek"):
                file_path_or_buffer.seek(0)

        wb = openpyxl.load_workbook(file_path_or_buffer, read_only=True, data_only=True)
        # Verifica que la hoja exista
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"Worksheet '{sheet_name}' does not exist. Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        fixed_start = column_index_from_string(fixed_range[0])
        fixed_end = column_index_from_string(fixed_range[1])

        # Definir dónde inicia el grupo de fechas según la hoja:
        if sheet_name.upper() == "US":
            date_start = column_index_from_string("S")
        elif sheet_name.upper() == "MN":
            date_start = column_index_from_string("N")
        else:
            date_start = fixed_end + 1

        end_month = None
        raw_date_cols = []
        # Iterar desde date_start hasta el final o hasta encontrar el stop_marker en la fila 2.
        for col_idx in range(date_start, ws.max_column + 1):
            cell_value = ws.cell(row=2, column=col_idx).value
            # Si se encuentra el stop_marker, se finaliza el grupo de fechas
            if cell_value is not None and cell_value == stop_marker:
                end_month = col_idx - 1
                break
            cell1 = ws.cell(row=1, column=col_idx).value
            combined = ""
            if cell1:
                combined += str(cell1).strip()
            if cell_value:
                combined += str(cell_value).strip()
            raw_date_cols.append(combined)
        if end_month is None:
            end_month = ws.max_column

        # Obtener las letras correspondientes al rango de fechas
        date_cols_letters = [
            get_column_letter(i) for i in range(date_start, end_month + 1)
        ]
        # Obtener las letras para el rango fijo
        fixed_cols = [get_column_letter(i) for i in range(fixed_start, fixed_end + 1)]
        usecols_str = ",".join(fixed_cols + date_cols_letters)

        # Procesar raw_date_cols para obtener los nombres de fechas
        # Filtrar solo aquellos que cumplan el patrón: 4 dígitos seguidos de al menos 3 letras
        raw_filtered = [
            col
            for col in raw_date_cols
            if re.match(r"^\d{4}[A-Za-z]{3,}$", col, re.IGNORECASE)
        ]
        # Convertir al formato deseado, ej.: "2021AGO" -> "agosto-2021"
        converted = [convert_date_col(col) for col in raw_filtered]
        # Ordenar cronológicamente
        converted = reorder_date_columns(converted)

        return usecols_str, converted

    def process_excel_files(
        self,
        file_path_or_buffer,
        pen_end_date: str = "2025-07-31",
        usd_end_date: str = "2025-12-31",
        read_type: str | None = None,
    ):
        """Procesar archivos Excel - LÓGICA EXACTA DiferidoExternoCalcular V1"""
        # PARA PEN:
        # Se obtiene el usecols y las columnas de fecha a partir del rango fijo ("B", "H")
        pen_usecols, pen_date_cols = self.auto_get_usecols(
            file_path_or_buffer,
            sheet_name="MN",
            fixed_range=("B", "H"),
            stop_marker="VAL",
        )
        print(f"Usecols PEN: {pen_usecols}")
        df_pen = self.read_excel_file(
            file_path_or_buffer,
            sheet_name="MN",
            usecols=pen_usecols,
            date_col="Fecha de Op",
        )
        print(
            f"dataframe df_pen: {df_pen.loc[df_pen['LIQUIDACIÓN'] == 'LIQ2502000076', :] if 'LIQUIDACIÓN' in df_pen.columns and not df_pen.empty else 'No data'}"
        )
        print(f"Columnas PEN: {pen_date_cols}")
        # Calcular el número de columnas fijas en PEN (de "B" a "H")
        fixed_pen_count = (
            column_index_from_string("H") - column_index_from_string("B") + 1
        )
        print(f"Columnas fijas PEN: {fixed_pen_count}")
        fixed_pen = list(df_pen.columns[:fixed_pen_count])
        print(f"Columnas fijas PEN: {fixed_pen}")
        expected_pen_cols = fixed_pen + pen_date_cols
        df_pen = df_pen.iloc[:, : len(expected_pen_cols)]
        df_pen.columns = expected_pen_cols
        df_pen["Moneda"] = "PEN"
        print(
            f"dataframe df_pen final: {df_pen.loc[df_pen['LIQUIDACIÓN'] == 'LIQ2502000076', :] if 'LIQUIDACIÓN' in df_pen.columns and not df_pen.empty else 'No data'}"
        )

        # PARA USD:
        # Se obtiene el usecols y las columnas de fecha a partir del rango fijo ("B", "G")
        usd_usecols, usd_date_cols = self.auto_get_usecols(
            file_path_or_buffer,
            sheet_name="US",
            fixed_range=("B", "G"),
            stop_marker="VAL",
        )

        df_usd = self.read_excel_file(
            file_path_or_buffer,
            sheet_name="US",
            usecols=usd_usecols,
            date_col="Fecha de Op",
        )

        # Calcular el número de columnas fijas en USD (de "B" a "G")
        fixed_usd_count = (
            column_index_from_string("G") - column_index_from_string("B") + 1
        )
        fixed_usd = list(df_usd.columns[:fixed_usd_count])
        print(f"Columnas fijas USD: {fixed_usd}")
        expected_usd_cols = fixed_usd + usd_date_cols
        df_usd = df_usd.iloc[:, : len(expected_usd_cols)]
        df_usd.columns = expected_usd_cols
        df_usd["Moneda"] = "USD"

        df = pd.concat([df_pen, df_usd], axis=0).rename(
            columns={
                "LIQUIDACIÓN": "CodigoLiquidacion",
                "Fecha de Op": "FechaOperacion",
                "Días Efect": "DiasEfectivo",
                "Intereses sin IGV": "Interes",
                "F.Pago Confirmada / Real": "FechaConfirmado",
            }
        )
        df["CodigoLiquidacion"] = df["CodigoLiquidacion"].str.strip()
        df["N° de Factura referencial"] = df["N° de Factura referencial"].str.strip()
        df.rename(columns={"N° de Factura referencial": "NroDocumento"}, inplace=True)
        df = df.drop(columns=["TIPO DE OPERACIÓN"], errors="ignore")
        return df

    def replace_columns_with_dates(self, df: pd.DataFrame):
        """Reemplazar columnas con fechas - LÓGICA EXACTA DiferidoExternoCalcular V1"""
        # Columnas fijas deseadas en el orden correcto
        columns_to_keep = [
            "CodigoLiquidacion",
            "NroDocumento",
            "FechaOperacion",
            "FechaConfirmado",
            "Moneda",
            "Interes",
            "DiasEfectivo",
        ]
        # Excluir las columnas indeseadas
        unwanted = ["Día de fecha de Op.", "Día de fecha de Pago"]
        # Tomar el resto de columnas (asumidas de fecha) sin las que no deseamos
        date_columns = [
            col
            for col in df.columns
            if col not in columns_to_keep and col not in unwanted
        ]
        date_columns = reorder_date_columns(date_columns)
        new_columns = columns_to_keep + date_columns
        return df[new_columns]

    def calcular_diferido_externo(
        self, file_buffer: Union[BytesIO, str], hasta: str
    ) -> pd.DataFrame:
        """
        Calcula el diferido externo - LÓGICA EXACTA DiferidoExternoCalcular V1

        Parámetros:
            file_buffer: Archivo Excel o buffer
            hasta (str): Cadena obligatoria con el formato "YYYY-MM"

        Retorna:
            pd.DataFrame: DataFrame final procesado
        """
        # Validamos el formato
        if not hasta or not re.match(r"^\d{4}-\d{2}$", hasta):
            raise ValueError(
                "El parámetro 'hasta' es obligatorio y debe tener el formato 'YYYY-MM'."
            )

        # Procesar el archivo Excel y obtener el DataFrame con los datos mensuales
        df = self.process_excel_files(file_buffer)

        # Convertir 'hasta' a fecha y obtener su último día
        hasta_date = datetime.strptime(hasta, "%Y-%m")
        hasta_date = self.last_day_of_month(hasta_date)

        # Filtrar según FechaOperacion
        df = df[df["FechaOperacion"] <= hasta_date]

        # Reorganizar columnas y limpiar DataFrame
        df = self.replace_columns_with_dates(df)
        df = df.dropna(subset=["CodigoLiquidacion"])
        df = df.fillna(0)
        return df

    # ==================== COMPARAR DIFERIDOS - LÓGICA EXACTA V1 ====================

    def comparar_diferidos(
        self, df_externo: pd.DataFrame, df_calculado: pd.DataFrame
    ) -> pd.DataFrame:
        """
        Compara diferidos - LÓGICA EXACTA DiferidoCalcular V1
        """
        # Identificar las columnas de fecha (con el patrón "Mes-YYYY")
        date_cols_externo = set(
            col for col in df_externo.columns if re.match(r"^[A-Za-z]+-\d{4}$", col)
        )
        date_cols_calculado = set(
            col for col in df_calculado.columns if re.match(r"^[A-Za-z]+-\d{4}$", col)
        )
        # Unión de ambas y ordenar cronológicamente
        union_date_cols = list(date_cols_externo.union(date_cols_calculado))
        date_columns = self.reorder_date_columns(union_date_cols)

        # Reordenamos las columnas fijas deseadas (nota: "CodigoLiquidacion" y "Moneda" son índices)
        # Queremos fijar el orden: NroDocumento, Interes, FechaOperacion, FechaConfirmado, DiasEfectivo
        fixed_cols = [
            "CodigoLiquidacion",
            "NroDocumento",
            "Interes",
            "FechaOperacion",
            "FechaConfirmado",
            "Moneda",
            "DiasEfectivo",
        ]
        # Reindexar ambos DataFrames para tener los fijos y las columnas de fechas en el mismo orden.
        df_externo = df_externo.reindex(columns=fixed_cols + date_columns, fill_value=0)
        df_calculado = df_calculado.reindex(
            columns=fixed_cols + date_columns, fill_value=0
        )

        # Definir las columnas numéricas a agrupar (todo excepto las de identificación y las fechas)
        numeric_cols = [
            col
            for col in df_externo.columns
            if col
            not in ["CodigoLiquidacion", "Moneda", "FechaOperacion", "FechaConfirmado"]
        ]

        # Agrupación numérica: suma (solo se suman columnas numéricas)
        df_numeric_grouped = df_externo.groupby(["CodigoLiquidacion", "Moneda"])[
            numeric_cols
        ].sum(numeric_only=True)
        calc_numeric_grouped = df_calculado.groupby(["CodigoLiquidacion", "Moneda"])[
            numeric_cols
        ].sum(numeric_only=True)

        # Agrupación de las columnas fijas no numéricas: usamos "last"
        df_fixed_grouped = df_externo.groupby(["CodigoLiquidacion", "Moneda"])[
            ["NroDocumento", "FechaOperacion", "FechaConfirmado"]
        ].agg("last")
        calc_fixed_grouped = df_calculado.groupby(["CodigoLiquidacion", "Moneda"])[
            ["NroDocumento", "FechaOperacion", "FechaConfirmado"]
        ].agg("last")

        # Agrupación de las columnas que no estuvieron en la fija y que se quieran mantener (por ejemplo, DiasEfectivo e Interes)
        # Notar que "Interes" y "DiasEfectivo" aparecen en la suma; de allí se tomarán sus valores sumados.
        # Unir ambas agregaciones:
        df_grouped = df_numeric_grouped.join(df_fixed_grouped)
        calc_grouped = calc_numeric_grouped.join(calc_fixed_grouped)

        # Combinar en un único DataFrame usando keys "Externo" y "Calculado"
        sums = pd.concat(
            [df_grouped, calc_grouped], axis=1, keys=["Externo", "Calculado"]
        )

        # Calcular diferencias para cada columna de fechas (de meses)
        for col in date_columns:
            sums[("Diferencia_Monto", col)] = (
                sums[("Externo", col)] - sums[("Calculado", col)]
            )

        diff_cols = [("Diferencia_Monto", col) for col in date_columns]
        sums["Diferencia_Significativa"] = sums[diff_cols].abs().ge(1).any(axis=1)

        # Reordenar las columnas en cada nivel para que en "Externo" y "Calculado" aparezcan primero
        # las columnas fijas ordenadas según lo deseado, seguidas de las columnas de fechas.
        fijo_orden = [
            "NroDocumento",
            "Interes",
            "FechaOperacion",
            "FechaConfirmado",
            "DiasEfectivo",
        ]

        def reordenar_level(key):
            # Para un nivel dado (Externo o Calculado), seleccionar columnas fijas según el orden definido
            fixed_level = [(key, col) for col in fijo_orden]
            date_level = [(key, col) for col in date_columns]
            return fixed_level + date_level

        nuevos_cols = (
            reordenar_level("Externo")
            + reordenar_level("Calculado")
            + [("Diferencia_Monto", col) for col in date_columns]
            + [("Diferencia_Significativa", "")]
        )
        sums = sums[nuevos_cols]
        return sums

    # ==================== MÉTODOS PRINCIPALES - COMPATIBILIDAD TOTAL V1 ====================

    def calcular_diferido(
        self, file_buffer: Union[BytesIO, str], df_interno: pd.DataFrame, hasta: str
    ) -> pd.DataFrame:
        """
        Calcula diferido completo - LÓGICA EXACTA DiferidoCalcular V1
        """
        if not hasta or not re.match(r"^\d{4}-\d{2}$", hasta):
            raise ValueError(
                "El parámetro 'hasta' es obligatorio y debe tener el formato 'YYYY-MM'."
            )
        # Calcular el diferido interno
        diferido_interno_df = self.calcular_diferido_interno(df_interno, hasta=hasta)
        # Calcular el diferido externo
        diferido_externo_df = self.calcular_diferido_externo(file_buffer, hasta=hasta)
        return self.comparar_diferidos(diferido_externo_df, diferido_interno_df)

    async def calcular_diferido_async(
        self, file_buffer: Union[BytesIO, str], df_interno: pd.DataFrame, hasta: str
    ) -> pd.DataFrame:
        """
        Calcula diferido completo asíncrono - LÓGICA EXACTA DiferidoCalcular V1
        """
        if not hasta or not re.match(r"^\d{4}-\d{2}$", hasta):
            raise ValueError(
                "El parámetro 'hasta' es obligatorio y debe tener el formato 'YYYY-MM'."
            )
        # Ejecutar métodos pesados en hilos separados
        diferido_interno_df = await asyncio.to_thread(
            self.calcular_diferido_interno, df_interno, hasta
        )
        diferido_externo_df = await asyncio.to_thread(
            self.calcular_diferido_externo, file_buffer, hasta
        )
        resultado = await asyncio.to_thread(
            self.comparar_diferidos, diferido_externo_df, diferido_interno_df
        )
        return resultado

    async def reorder_date_columns_async(self, date_cols: list[str]) -> list[str]:
        """Versión asíncrona de reorder_date_columns - COMPATIBILIDAD V1"""
        # Ejecuta la función sincrónica en un hilo separado
        return await asyncio.to_thread(self.reorder_date_columns, date_cols)

    async def comparar_diferidos_async(self, df_externo, df_calculado):
        """Versión asíncrona de comparar_diferidos - COMPATIBILIDAD V1"""
        # Ejecuta la función comparar_diferidos en un hilo separado
        return await asyncio.to_thread(
            self.comparar_diferidos, df_externo, df_calculado
        )
